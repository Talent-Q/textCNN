# -*- coding: utf-8 -*-#
# -------------------------------------------------------------------------------
#
# FileName:     tools
# Author:       TalentQ
# Date:         2023-8-31
#
# -------------------------------------------------------------------------------
import json
import random
import openpyxl
from typing import Dict, List


def addr_related_txs_num_To_dict(path1: str, path2: str) -> Dict:
    addresses = {}
    with open(path1, 'r') as f:
        lines = f.readlines()
        for line in lines:
            line = line.rstrip()
            res = line.split()
            addresses[res[0]] = res[1]
    with open(path2, 'r') as f:
        lines = f.readlines()
        for line in lines:
            line = line.rstrip()
            res = line.split()
            addresses[res[0]] = res[1]
    return addresses


def read_txs_json(file: str) -> List:
    with open(file, 'r', encoding="utf-8") as f:
        lines = f.readlines()
        for i, line in enumerate(lines):
            lines[i] = json.loads(line)
    return lines


def shuffle_dataset(dataset: List, label: List):
    tmp = list(zip(dataset, label))
    random.shuffle(tmp)

    return zip(*tmp)


def _writeinexcel2(datalist, sheet):
    for z in datalist:
        sheet.append([z])


def _writelabel(label, sheet):
    m = 1
    n = 2
    for i in label:
        sheet.cell(row=m, column=n, value=i)
        m = m + 1


def back_up_dataset(dataset, label, to_path):
    wb = openpyxl.Workbook()  # ,encoding = 'utf-8'
    sheet = wb.create_sheet('dataset', 0)
    _writeinexcel2(dataset, sheet)
    _writelabel(label, sheet)
    wb.save(to_path)


def random_data(sheet):
    numbers = list(range(1, sheet.max_row + 1))
    random.shuffle(numbers)
    for i in numbers:
        row = sheet[i]
        r = []
        for cell in row:
            r.append(cell.value)
        sheet.append(r)
    sheet.delete_rows(1, sheet.max_row // 2)


def trainset(sheetfrom, sheetto, start, end):
    max_col = sheetfrom.max_column
    # print(max_col)
    for x in range(start, end + 1):
        row_data = []
        for y in range(1, max_col + 1):
            cell_value = sheetfrom.cell(row=x, column=y).value
            row_data.append(cell_value)
        sheetto.append(row_data)
