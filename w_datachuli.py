import json
import os
import openpyxl
import random
from keras.preprocessing.text import Tokenizer, text_to_word_sequence
from util.tools import str_to_json


def tongyigeshi(rows, normal_addr_txs_num_path, special_addr_txs_num_path):  # 输入json,统一格式
    addresses = str_to_json(normal_addr_txs_num_path, special_addr_txs_num_path)
    maxlength = 0
    rowslist = []
    for row in rows:
        tx_hash = row.get("hash")
        rowinput = row.get("inputs")
        rowoutput = row.get("outputs")
        inputad = []
        inputvalue = []
        inputscript = []
        input_prehash = []
        outputad = []
        outputvalue = []
        outputscript = []
        outputtype = []
        outputdata_string = []
        addr_txs_num = []
        for a in rowinput:
            inputad.append(a.get("addresses"))
            inputvalue.append(a.get("output_value"))
            inputscript.append(a.get("script"))
            input_prehash.append(a.get("prev_hash"))
            for i in a.get("addresses"):
                addr_txs_num.append(addresses[i])
        for a in rowoutput:
            outputad.append(a.get("addresses"))
            outputvalue.append(a.get("value"))
            outputscript.append(a.get("script"))
            outputtype.append(a.get("script_type"))
            outputdata_string.append(a.get("data_hex"))  # data_string
            if a.get("addresses") is not None:
                for i in a.get("addresses"):
                    addr_txs_num.append(addresses[i])

        # 6. 调整了 原始 的字段顺序，并删除了几个字段, 并加入 addr_txs_num (0.926, 0.959, 0.933)
        # rowlist = [row.get("total"),    # number
        #            row.get("fees"),     # number
        #            row.get("vin_sz"),   # number
        #            inputvalue,          # number
        #            row.get("vout_sz"),  # number
        #            outputvalue,         # number
        #            addr_txs_num,        # number
        #            inputad,
        #            outputad,
        #            inputscript,
        #            outputscript]

        # 5. 调整了 原始 的字段顺序，并加入 addr_txs_num (0.945, 0.952, 0.911, 0.947, 0.926) => 0.94
        rowlist = [row.get("total"),    # number
                   row.get("fees"),     # number
                   row.get("vin_sz"),   # number
                   inputvalue,          # number
                   row.get("vout_sz"),  # number
                   outputvalue,         # number
                   addr_txs_num,        # number
                   inputad,
                   outputad,
                   tx_hash,
                   input_prehash,
                   inputscript,
                   outputscript,
                   outputtype,
                   outputdata_string]

        # 4. 原始， 加入 addr_txs_num （0.935, 0.921, 0.971, 0.94, 0.945）=> 0.94
        # rowlist = [row.get("total"),  # number
        #            row.get("fees"),  # number
        #            row.get("vin_sz"),  # number
        #            addr_txs_num,  # number，加入的字段
        #            inputad,
        #            inputvalue,  # number
        #            inputscript,
        #            row.get("vout_sz"),  # number
        #            outputad,
        #            outputvalue,  # number
        #            outputscript,
        #            outputtype,
        #            outputdata_string,
        #            tx_hash,
        #            input_prehash]

        # 3. 调整了 原始 的字段顺序，并删除了几个字段 (0.861, 0.778, 0.816, 0.825, 0.84) => 0.827
        # rowlist = [row.get("total"),    # number
        #            row.get("fees"),     # number
        #            row.get("vin_sz"),   # number
        #            inputvalue,          # number
        #            row.get("vout_sz"),  # number
        #            outputvalue,         # number
        #            inputad,
        #            outputad,
        #            inputscript,
        #            outputscript]

        # 2. 调整了 原始 的字段顺序 (0.804, 0.785, 0.792, 0.773, 0.847) => 0.794
        # rowlist = [row.get("total"),    # number
        #            row.get("fees"),     # number
        #            row.get("vin_sz"),   # number
        #            inputvalue,          # number
        #            row.get("vout_sz"),  # number
        #            outputvalue,         # number
        #            inputad,
        #            outputad,
        #            tx_hash,
        #            input_prehash,
        #            inputscript,
        #            outputscript,
        #            outputtype,
        #            outputdata_string]

        # 1. 原始 (0.761, 0.684,  0.77, 0.768, 0.763) => 0.764
        # rowlist = [row.get("total"),  # number
        #            row.get("fees"),  # number
        #            row.get("vin_sz"),  # number
        #            inputad,
        #            inputvalue,  # number
        #            inputscript,
        #            row.get("vout_sz"),  # number
        #            outputad,
        #            outputvalue,  # number
        #            outputscript,
        #            outputtype,
        #            outputdata_string,
        #            tx_hash,
        #            input_prehash]

        rowlist = str(rowlist)

        rowlist = rowlist.replace('\\n', '')
        rowlist = rowlist.replace('\"', '')
        rowlist = rowlist.replace('\'', '')
        rowlist = rowlist.replace('{', '')
        rowlist = rowlist.replace('}', '')
        rowlist = rowlist.replace('[', '')
        rowlist = rowlist.replace(']', '')
        rowlist = rowlist.replace(' ', '')

        rowlist = " ".join(rowlist)
        length = len(rowlist)
        if length > maxlength:
            maxlength = length

        rowslist.append(rowlist)
    print("#maxlength=", maxlength)
    return rowslist

    # "total": 98992600,
    # "fees": 7400,
    #
    # "vin_sz": 1,
    # "inputaddressed": ["1Phsizjg38ZMju6EdgzuBDc8keNq7gdgub"],
    # "inputvalue": ["99000000"],
    # "inputscript":"47304402200dd69d1300bf612abf09130d1d2e39073aa9a724fbb80c65846b14bfc8e628ea02204f4da818e35b734743494232507978e7b10ba067fd3043d50ba957fb78dafa20014104bd7ff41dcc3cf34c9feac1beb48ce1844dc1faee3432f2800c0b6c8214cd90e4e4d71a5594c74964dc2234f69fe854494a199ca7d8dc4cec0778d12073ae8e8c",
    #
    # "vout_sz": 3,
    # "outputaddressed":[NULL,"16Wnx5y6V3yjEFHrCrGPnbAZ4fscA8AhAo","1FwZt9qtRVXxS1wNbuBF1AhVdf2GFrtCBX"],
    # "outputvalue":[0,98654723,337877,],
    # "outputscript":["6a4c504441343144434143463735453543443536323344384339333839443642453435343734383243464542343643393537423046333638464336313038303536393337343635373337343331333133313045","76a9143c7b0c79baf8111a251e60af4a27b0b705e4710788ac","76a914a3e3abdee5f8cb30cc556a9a70d763905723390688ac",
    #                ]
    # "outputdata_string":["DA41DCACF75E5CD5623D8C9389D6BE4547482CFEB46C957B0F368FC610805693746573743131310E",NULL,NULL]


def tiquziduan(rows):
    maxlength = 0
    rowslist = []
    for row in rows:
        rowinput = row.get("inputs")
        rowoutput = row.get("outputs")
        inputad = []
        inputvalue = []
        inputscript = []
        outputad = []
        outputvalue = []
        outputscript = []
        outputdata_string = []
        for a in rowinput:
            inputad.append(a.get("addresses"))
            inputvalue.append(a.get("output_value"))
            inputscript.append(a.get("script"))
        for a in rowoutput:
            outputad.append(a.get("addresses"))
            outputvalue.append(a.get("value"))
            outputscript.append(a.get("script"))
            outputdata_string.append(a.get("data_hex"))  # data_string

        rowlist = []
        # rowlist.append(row.get("total"))
        # rowlist.append(row.get("fees"))
        # rowlist.append(row.get("vin_sz"))
        # rowlist.append(inputad)
        # rowlist.append(inputvalue)
        rowlist.append(inputscript)
        # rowlist.append(row.get("vout_sz"))
        # rowlist.append(outputad)
        # rowlist.append(outputvalue)
        # rowlist.append(outputscript)
        # rowlist.append(outputdata_string)
        rowlist = str(rowlist)

        rowlist = rowlist.replace('\\n', '')
        rowlist = rowlist.replace('\"', '')
        rowlist = rowlist.replace('\'', '')
        rowlist = rowlist.replace('{', '')
        rowlist = rowlist.replace('}', '')
        rowlist = rowlist.replace('[', '')
        rowlist = rowlist.replace(']', '')
        rowlist = rowlist.replace(' ', '')

        rowlist = " ".join(rowlist)
        length = len(rowlist)
        if length > maxlength:
            maxlength = length

        rowslist.append(rowlist)
    print("#maxlength=", maxlength)
    return rowslist


def openfile(file1):
    rows = []
    file = open(file1, 'r', encoding="utf-8")  # 读取的文件
    for line in file.readlines():
        line = json.loads(line)
        rows.append(line)
    return rows


def writeinexcel2(datalist, sheet):
    for z in datalist:
        sheet.append([z])


def writelabel(label, sheet):
    m = 1
    n = 2
    for i in label:
        sheet.cell(row=m, column=n, value=i)
        m = m + 1


def random_data(sheet):
    numbers = list(range(1, sheet.max_row + 1))
    random.shuffle(numbers)
    for i in numbers:
        row = sheet[i]
        r = []
        for cell in row:
            r.append(cell.value)
        sheet.append(r)
    sheet.delete_rows(1, sheet.max_row // 2)


def trainset(sheetfrom, sheetto, start, end):
    max_col = sheetfrom.max_column
    # print(max_col)
    for x in range(start, end + 1):
        row_data = []
        for y in range(1, max_col + 1):
            cell_value = sheetfrom.cell(row=x, column=y).value
            row_data.append(cell_value)
        sheetto.append(row_data)


def ciqianru(alldata, label):
    tokenizer = Tokenizer(lower=False, num_words=100)  # 构建索引字典
    tokenizer.fit_on_texts(alldata)  # 构建索引单词
    word_index = tokenizer.word_index
    print("#word_index", word_index)

    sequences = tokenizer.texts_to_sequences(alldata)  # 将字符串转换为整数索引组成的列表
    # print("type(sequences)",type(sequences),len(sequences))
    data_all = []
    for string in sequences:
        data_all.append(string)

    return data_all, label


def shujuchuli(normalpath="./dataset/1in2outPN.json", 
                specialpath="./dataset/7_CHT02.json", 
                normal_addr_txs_num_path='./dataset/data_ql/1in2outPN_addr_num.txt', 
                special_addr_txs_num_path='./dataset/data_ql/7_CHT02_addr_num.txt'):
    object = 'all'  # 'ziduan' or 'all'
    print("数据提取对象:", object)

    normalpath = normalpath  # normal.json
    normal = openfile(normalpath)
    if object == 'all':
        normaldata = tongyigeshi(normal, normal_addr_txs_num_path, special_addr_txs_num_path)
    else:
        normaldata = tiquziduan(normal)

    specialpath = specialpath  # special.json
    special = openfile(specialpath)
    if object == 'all':
        specialdata = tongyigeshi(special, normal_addr_txs_num_path, special_addr_txs_num_path)
    else:
        specialdata = tiquziduan(special)

    print("normal数据:", normaldata[4])
    print("special数据:", specialdata[4])
    print("normal数据:", normalpath)
    print("special数据:", specialpath)

    # print("########### 备份混合数据集 ##############################")
    alldata = normaldata + specialdata  # list
    label = []
    for i in range(len(normaldata)):
        label.append(0)
    for i in range(len(specialdata)):
        label.append(1)

    c = list(zip(alldata, label))
    random.shuffle(c)
    alldata, label = zip(*c)

    file3 = openpyxl.Workbook()  # ,encoding = 'utf-8'
    sheet3 = file3.create_sheet('dataset', 0)
    writeinexcel2(alldata, sheet3)
    writelabel(label, sheet3)
    file3.save('./dataset/file/dataset.xlsx')

    # print("########### 转变为词向量 ##############################")

    alldata_t, label_t = ciqianru(alldata, label)

    return alldata_t, label_t, normalpath, specialpath


if __name__ == '__main__':
    # shujuchuli(normalpath="./dataset/1in2outPN.json", 
    #             specialpath="./dataset/7_CHT02.json",
    #             normal_addr_txs_num_path='./dataset/data_ql/1in2outPN_addr_num.txt', 
    #             special_addr_txs_num_path='./dataset/data_ql/7_CHT02_addr_num.txt')
    
    shujuchuli(normalpath="./dataset/1in2outPN.json", 
                specialpath="./dataset/6_DSA.json",
                normal_addr_txs_num_path='./dataset/data_ql/1in2outPN_addr_num.txt', 
                special_addr_txs_num_path='./dataset/data_ql/6_DSA_addr_num.txt')
